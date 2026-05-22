"""
NORAD Discovery Engine — Backend & API Cost Workbook
Builds NORAD_COST_REPORT.xlsx — a SaaS-grade, stakeholder-ready cost model.

8 tabs:
  1. Executive Summary
  2. Tool Catalog (master)
  3. Per-Candidate Cost Flow
  4. Workflow Scenarios
  5. Volume Sensitivity
  6. LLM Models Reference
  7. Database & Vector Storage Comparison
  8. Assumptions & Notes

All vendor pricing verified April–May 2026.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ============================================================
# DESIGN SYSTEM
# ============================================================
NAVY        = "0F2A4A"
NAVY_LIGHT  = "1E4675"
GREY_LIGHT  = "F4F6F9"
GREY_MED    = "E6EEF5"
ACCENT_GREEN= "1F6B3A"
ACCENT_RED  = "B23A3A"
ACCENT_AMBR = "B58400"
WHITE       = "FFFFFF"

THIN  = Side(style="thin",  color="BFC9D4")
MED   = Side(style="medium", color=NAVY)

def box(top=THIN, bot=THIN, lef=THIN, rig=THIN):
    return Border(top=top, bottom=bot, left=lef, right=rig)

def fill(hex_):
    return PatternFill("solid", fgColor=hex_)

# Common fonts
F_TITLE   = Font(name="Calibri", size=18, bold=True, color=NAVY)
F_SUBTTL  = Font(name="Calibri", size=11, italic=True, color="555555")
F_HEAD    = Font(name="Calibri", size=10, bold=True, color=WHITE)
F_SUBHEAD = Font(name="Calibri", size=10, bold=True, color=NAVY)
F_BODY    = Font(name="Calibri", size=10, color="222222")
F_BODY_B  = Font(name="Calibri", size=10, bold=True, color="222222")
F_TOTAL   = Font(name="Calibri", size=11, bold=True, color=WHITE)
F_KPI_LBL = Font(name="Calibri", size=10, bold=True, color="555555")
F_KPI_VAL = Font(name="Calibri", size=20, bold=True, color=NAVY)
F_NOTE    = Font(name="Calibri", size=9, italic=True, color="666666")
F_LINK    = Font(name="Calibri", size=9, color="0563C1", underline="single")

# Common fills
FILL_HEAD     = fill(NAVY)
FILL_SUBHEAD  = fill(GREY_MED)
FILL_TOTAL    = fill(NAVY_LIGHT)
FILL_ALT      = fill(GREY_LIGHT)
FILL_KPI      = fill("FFFFFF")
FILL_GOOD     = fill("DCEEDF")
FILL_WARN     = fill("FFF1D1")
FILL_BAD      = fill("FBDADA")

# Common alignments
A_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
A_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
A_RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=True)
A_TOP_L  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)


# ============================================================
# HELPERS
# ============================================================
def write_row(ws, row, values, *, font=F_BODY, fill_=None, align=A_LEFT,
              border=None, fmts=None):
    """Write a row of values starting at column A."""
    for i, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=val)
        c.font = font
        c.alignment = align
        if fill_:
            c.fill = fill_
        if border:
            c.border = border
        if fmts and len(fmts) > i - 1 and fmts[i - 1]:
            c.number_format = fmts[i - 1]


def header_row(ws, row, headers, *, height=32):
    ws.row_dimensions[row].height = height
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.alignment = A_CENTER
        c.border = box()


def title_block(ws, title, subtitle, last_col_letter="J"):
    ws.row_dimensions[1].height = 34
    ws.merge_cells(f"A1:{last_col_letter}1")
    c = ws["A1"]
    c.value = title
    c.font = F_TITLE
    c.alignment = A_LEFT

    ws.row_dimensions[2].height = 18
    ws.merge_cells(f"A2:{last_col_letter}2")
    c2 = ws["A2"]
    c2.value = subtitle
    c2.font = F_SUBTTL
    c2.alignment = A_LEFT
    ws.row_dimensions[3].height = 6


def section_label(ws, row, text, span_letter):
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"A{row}:{span_letter}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
    c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# WORKBOOK
# ============================================================
wb = Workbook()

# ============================================================
# TAB 1 — EXECUTIVE SUMMARY
# ============================================================
ws = wb.active
ws.title = "Executive Summary"
ws.sheet_properties.tabColor = NAVY
ws.sheet_view.showGridLines = False

set_col_widths(ws, [3, 24, 22, 22, 22, 22, 22, 3])

title_block(
    ws,
    "NORAD Discovery Engine — Backend & API Cost Report",
    "Verified vendor pricing · April–May 2026 · Premium SaaS-grade stack · Per-deployment economics",
    last_col_letter="H",
)

# KPI strip
section_label(ws, 4, "Headline economics — single NORAD deployment", "H")
ws.row_dimensions[5].height = 6

kpis = [
    ("v1 Monthly Cost",            "$3,042",  "Batch + dashboard"),
    ("Phase 2 Monthly Cost",       "$4,351",  "+ NLS + multi-tenant"),
    ("Annual v1",                  "$36,504", "12 × monthly"),
    ("Per Candidate Processed",    "$0.51",   "6,000 candidates/mo"),
    ("Per Qualified Opportunity",  "$38",     "≥70 score, ~80/mo"),
    ("Per Elite Alert",            "$254",    "≥85 score, ~12/mo"),
]

# Layout 6 KPI cards across cols B–G (3 wide × 2 rows)
def kpi_card(ws, row, col, label, value, sub):
    cell_label = ws.cell(row=row,     column=col, value=label)
    cell_val   = ws.cell(row=row + 1, column=col, value=value)
    cell_sub   = ws.cell(row=row + 2, column=col, value=sub)
    cell_label.font = F_KPI_LBL
    cell_val.font   = F_KPI_VAL
    cell_sub.font   = F_NOTE
    for c in (cell_label, cell_val, cell_sub):
        c.alignment = Alignment(horizontal="center", vertical="center")
    # Box around card
    for r in range(row, row + 3):
        for cc in range(col, col + 1):
            ws.cell(row=r, column=cc).border = box()
            ws.cell(row=r, column=cc).fill = FILL_KPI
    ws.row_dimensions[row].height = 18
    ws.row_dimensions[row + 1].height = 32
    ws.row_dimensions[row + 2].height = 18

# Row 1 of cards: cols B(2), D(4), F(6) — leave gap cols
positions = [(6, 2), (6, 4), (6, 6), (10, 2), (10, 4), (10, 6)]
for (label, value, sub), (r, c) in zip(kpis, positions):
    kpi_card(ws, r, c, label, value, sub)

# Highlight v1 + Phase 2 cells in green
ws.cell(row=7, column=2).fill = FILL_GOOD  # v1 monthly
ws.cell(row=7, column=4).fill = FILL_WARN  # Phase 2 monthly

# Volume baseline block
section_label(ws, 14, "Volume model (per deployment, per day)", "H")
ws.row_dimensions[15].height = 8

header_row(ws, 16, ["Funnel stage", "Daily", "Monthly", "What it means", "", "", "", ""])
funnel = [
    ("Raw signals ingested (filings, news, patents, social, web)",
     "~5,000", "~150,000", "Twice-daily ingestion across all configured sources"),
    ("De-duplicated to candidate companies",
     "200", "6,000", "After entity resolution (Splink). PRIMARY VOLUME LEVER."),
    ("Pass hard filters (vertical, geo, size, tags)",
     "~60", "~1,800", "LLM justification budget tier"),
    ("Score ≥ 70 — review queue",
     "~3", "~80", "What an operator actually reviews"),
    ("Score ≥ 85 — Elite alert",
     "~0.4", "~12", "Top-of-funnel opportunities for client deal teams"),
]
for i, row in enumerate(funnel, start=17):
    fill_ = FILL_ALT if i % 2 else None
    bold = (i == 18)  # highlight the candidate row
    write_row(ws, i,
              [row[0], row[1], row[2], row[3], "", "", "", ""],
              font=F_BODY_B if bold else F_BODY,
              fill_=FILL_GOOD if bold else fill_,
              align=A_LEFT, border=box())
    ws.row_dimensions[i].height = 22

# v1 vs Phase 2 totals breakdown
section_label(ws, 24, "Monthly cost breakdown — v1 vs Phase 2 side-by-side", "H")
ws.row_dimensions[25].height = 8

header_row(ws, 26, ["Bucket", "v1 (batch + dashboard)", "Phase 2 (+NLS + multi-tenant)", "Delta", "", "", "", ""])
breakdown = [
    ("External Data APIs",     1181, 1930, 749),
    ("LLM Inference",            77,  187, 110),
    ("Core Infrastructure",   1784, 2234, 450),
]
r = 27
for label, v1, p2, dlt in breakdown:
    fill_ = FILL_ALT if r % 2 else None
    write_row(ws, r,
              [label, v1, p2, dlt, "", "", "", ""],
              font=F_BODY, fill_=fill_, align=A_LEFT, border=box(),
              fmts=[None, '"$"#,##0', '"$"#,##0', '"$"#,##0'])
    ws.row_dimensions[r].height = 20
    r += 1

# Total row
write_row(ws, r,
          ["TOTAL MONTHLY", 3042, 4351, 1309, "", "", "", ""],
          font=F_TOTAL, fill_=FILL_TOTAL, align=A_LEFT, border=box(),
          fmts=[None, '"$"#,##0', '"$"#,##0', '"$"#,##0'])
ws.row_dimensions[r].height = 26

# Footnote
r += 2
ws.merge_cells(f"A{r}:H{r}")
c = ws.cell(row=r, column=1,
            value="All numbers per single NORAD deployment at recommended baseline volume of 200 candidates/day. "
                  "Linear scaling: double the volume → ~2× external APIs + ~30% infra. See Volume Sensitivity tab.")
c.font = F_NOTE
c.alignment = A_LEFT

# Freeze panes and view
ws.freeze_panes = "A4"
ws.sheet_view.zoomScale = 100


# ============================================================
# TAB 2 — TOOL CATALOG (MASTER)
# ============================================================
ws = wb.create_sheet("Tool Catalog")
ws.sheet_properties.tabColor = NAVY_LIGHT
ws.sheet_view.showGridLines = False

# Columns:
# A # | B Layer | C Service | D What it does | E Plan/Tier | F Monthly Sub | G Unit Price |
# H Per 1K | I Per 10K | J Per 100K | K Rate Limit | L Pipeline Usage | M Vol/mo |
# N v1 $/mo | O Phase 2 $/mo | P Verdict / Why | Q Alternative considered | R Source

set_col_widths(ws,
    [4, 12, 22, 30, 22, 12, 22, 12, 12, 12, 22, 28, 12, 12, 12, 38, 38, 28])

title_block(
    ws,
    "Tool Catalog — Every Service in the NORAD Stack",
    "30 services · Premium tier choices · Verified pricing · Alternatives & verdict reasoning included",
    last_col_letter="R",
)

headers = [
    "#", "Layer", "Service / API", "What it does in NORAD", "Plan / Tier",
    "Monthly\nsub", "Unit price",
    "Per 1K", "Per 10K", "Per 100K",
    "Rate limit", "Pipeline usage", "Vol / mo",
    "v1\n$/mo", "Phase 2\n$/mo",
    "Verdict — why this pick",
    "Alternative considered",
    "Source",
]
header_row(ws, 4, headers, height=42)
ws.row_dimensions[4].height = 42

# Tool catalog data
# (#, Layer, Service, What, Plan, Sub$, Unit, /1K, /10K, /100K, Rate limit, Usage, Vol, v1$, p2$, Verdict, Alt, Source)
catalog = [
    # ---------- L1 — Search & web data ----------
    (1, "L1", "EXA AI",
     "Semantic search + AI-ready content extraction. Used 5 searches + 10 contents per candidate.",
     "Pay-as-you-go", 0,
     "$7/1K search · $1/1K content",
     "$7", "$70", "$700",
     "Configurable 180ms–1s latency, generous concurrency",
     "5 searches + 10 contents per candidate",
     30000, 270, 316,
     "WINNER. Token-efficient highlights designed for AI agents — cleanest LLM-ready output on the market.",
     "Brave Search API ($3/1K) — cheaper but no AI-grade highlights, more parsing work downstream.",
     "https://exa.ai/pricing"),

    (2, "L1", "SerpApi",
     "Google Search API. Trade-press + recent-news queries (~50 patterns × twice daily).",
     "Production",
     150, "$10 / 1K searches",
     "$10", "$100", "$1,000",
     "1,000 searches/hour",
     "~3,000 searches/mo",
     3000, 150, 150,
     "WINNER. Most reliable Google scraper, full SERP feature parity, US/CA geo-targeting.",
     "Bright Data SERP ($1.80–3/1K) — slightly cheaper bulk but uses CPM billing & IP-based geos.",
     "https://serpapi.com/pricing"),

    (3, "L1", "Bright Data SERP",
     "Backup search + Web Unlocker for geo-locked Google + tough sites.",
     "Micro",
     10, "$1.80 / 1K (CPM)",
     "$1.80", "$18", "$180",
     "5,555 req/mo at Micro tier",
     "Fallback only when SerpApi rate-limits",
     500, 10, 10,
     "Backup tier. Critical for redundancy on Google + global geo-targeting (195 countries).",
     "ScraperAPI, ZenRows — comparable price, less mature SERP parsing.",
     "https://brightdata.com/pricing"),

    (4, "L1", "Firecrawl",
     "Open-web crawl + structured extract with LLM-defined schemas. Replaces 10+ custom scrapers.",
     "Standard",
     83, "$0.00083 / credit (1 credit = 1 page)",
     "$0.83", "$8.30", "$83",
     "100K credits/mo, ~50 concurrent",
     "~6K maps + 15K scrapes + 6K extracts",
     27000, 83, 83,
     "WINNER. The 'extract' endpoint with LLM schemas saves weeks of glue code per source added.",
     "Apify Web Scraper actor — cheaper per page but no schema-driven extract.",
     "https://www.firecrawl.dev/pricing"),

    # ---------- L2 — Structured & regulatory ----------
    (5, "L2", "USPTO PatentSearch / ODP",
     "US patent filings (post-March 2026 ODP migration). Innovation signal.",
     "Free",
     0, "Free",
     "$0", "$0", "$0",
     "45 req/min per app key",
     "~1,000 req/mo",
     1000, 0, 0,
     "WINNER. Authoritative US patent source. Free.",
     "USPTO bulk downloads — heavier ETL, no benefit at our volume.",
     "https://data.uspto.gov/apis/api-rate-limits"),

    (6, "L2", "Google Patents Public Data (BigQuery)",
     "Global patent families, statistical analysis, bulk patent intelligence.",
     "Free tier",
     0, "First 1 TiB/mo free, then $6.25/TiB",
     "n/a", "n/a", "n/a",
     "BigQuery query quotas",
     "<200 GB scanned/mo",
     0, 0, 0,
     "WINNER. Free at our scan volume. Best global patent depth available.",
     "Lens.org — richer metadata but $1K/yr commercial licence.",
     "https://github.com/google/patents-public-data"),

    (7, "L2", "EPO OPS / CIPO / WIPO",
     "European, Canadian, World patent offices.",
     "Free",
     0, "Free",
     "$0", "$0", "$0",
     "EPO: 4GB/week. CIPO: standard.",
     "~500 req/mo combined",
     500, 0, 0,
     "Standard included. Covers non-US patent geographies for free.",
     "Lens.org consolidates all of these — but at $83/mo amortized.",
     "https://ops.epo.org/"),

    (8, "L2", "EDGAR / SEDAR+ / Companies House",
     "US, Canadian, UK corporate filings.",
     "Free",
     0, "Free",
     "$0", "$0", "$0",
     "Standard rate limits, no key for most",
     "~2,000 req/mo combined",
     2000, 0, 0,
     "Standard included. Authoritative public-company + UK private filings.",
     "Refinitiv / S&P paid feeds — overkill at this stage.",
     "https://www.sec.gov/edgar/searchedgar/companysearch.html"),

    (9, "L2", "OpenFDA / EMA / Health Canada / MHRA",
     "Health regulatory submissions (drug, device, supplement).",
     "Free",
     0, "Free",
     "$0", "$0", "$0",
     "Standard public API limits",
     "~3,000 req/mo combined",
     3000, 0, 0,
     "Standard included. Authoritative health-vertical signal.",
     "FDA bulk downloads — heavier ETL, no benefit at our volume.",
     "https://open.fda.gov/"),

    (10, "L2", "Lens.org Patent + Scholar API",
     "Richer patent metadata + academic-paper grounding (Phase 2 nice-to-have).",
     "Individual Commercial",
     0, "$1,000 / year (~$83/mo)",
     "n/a", "n/a", "n/a",
     "Standard tier limits",
     "Phase 2 only",
     0, 0, 83,
     "Phase 2 add-on. Adds academic + global patent depth above the free baseline.",
     "PatSnap — enterprise-grade but $10K+/yr.",
     "https://support.lens.org/knowledge-base/lens-patent-and-scholar-api/"),

    # ---------- L3 — Closed platforms ----------
    (11, "L3", "Apify — Scale plan",
     "Platform for all closed-platform actors. Includes $199 worth of compute credits.",
     "Scale",
     199, "$0.16 / compute unit",
     "n/a", "n/a", "n/a",
     "Per-actor concurrency",
     "Plan covers all actor compute",
     0, 199, 199,
     "WINNER. Scale tier gets Silver Store discount (~10%) + lower CU rate vs Starter.",
     "Business plan ($999) — only worth it above 4× our compute volume.",
     "https://apify.com/pricing"),

    (12, "L3", "Apify — LinkedIn Scraper",
     "Company + founder profiles. ~3 profiles per candidate.",
     "Per-result actor",
     0, "$3 / 1K results",
     "$3", "$30", "$300",
     "Bound by Apify plan concurrency",
     "~18,000 profiles/mo",
     18000, 54, 54,
     "WINNER. Most reliable cookie-less LinkedIn scraper on Apify Store.",
     "Phantombuster — comparable but per-launch fee model adds up faster.",
     "https://apify.com/get-leads/linkedin-scraper"),

    (13, "L3", "Apify — Instagram Scraper",
     "Brand presence + handle / engagement check.",
     "Per-result actor",
     0, "$1.60 / 1K results",
     "$1.60", "$16", "$160",
     "Bound by Apify plan",
     "~1,200 results/mo",
     1200, 2, 2,
     "WINNER. Apify-official actor, well-maintained.",
     "Bright Data Instagram dataset — bulk-only, expensive at low volume.",
     "https://apify.com/apify/instagram-scraper"),

    (14, "L3", "Apify — TikTok Scraper",
     "Pay-per-event scraper for video / profile / hashtag pulls.",
     "Per-event actor",
     0, "~$2 / 1K videos",
     "$2", "$20", "$200",
     "Bound by Apify plan",
     "~1,000 videos/mo",
     1000, 2, 2,
     "WINNER. clockworks/tiktok-scraper is the de facto standard.",
     "TikTok Research API — academic only, not commercial.",
     "https://apify.com/clockworks/tiktok-scraper"),

    (15, "L3", "Apify — PitchBook Scraper",
     "Funding history, cap table, competitor map for private companies.",
     "Per-result actor",
     0, "~$15 / 1K results",
     "$15", "$150", "$1,500",
     "Bound by Apify plan",
     "~1,800 profiles/mo",
     1800, 27, 27,
     "WINNER. Only practical way to programmatically access PitchBook short of full enterprise licence ($30K+/yr).",
     "Crunchbase Enterprise API — comparable data, $50K+/yr commitment.",
     "https://apify.com/store?search=pitchbook"),

    (16, "L3", "Apify — Reddit Scraper",
     "Subreddit + post + comment monitoring (community signal).",
     "CU-based actor",
     0, "Per-CU (bundled in Scale plan)",
     "n/a", "n/a", "n/a",
     "Bound by Apify plan",
     "~10K posts/mo",
     10000, 30, 30,
     "WINNER vs Reddit official API. Reddit's commercial API is contract-only and expensive ($10K+/yr).",
     "Reddit official API (commercial) — contract only, not viable at our scale.",
     "https://apify.com/store?search=reddit"),

    (17, "L3", "Grok Live Search (xAI)",
     "X (Twitter) + live web grounding. Replaces direct X API for most reads.",
     "Pay-per-token",
     0, "$1.25 in / $2.50 out per 1M tokens",
     "n/a", "n/a", "n/a",
     "Tier-based, generous",
     "~10K X/web grounding queries/mo",
     10000, 30, 50,
     "WINNER for X access. Built by xAI who own X — native deep access + per-token economics beat X API at our volume.",
     "X API direct — pay-per-read $0.005/post hits 2M cap fast and costs more.",
     "https://x.ai/api"),

    (18, "L3", "X (Twitter) API",
     "Minimal direct API budget for posting / specific endpoints Grok doesn't cover.",
     "Pay-per-use",
     0, "$0.01 post create · $0.005 post read",
     "$5", "$50", "$500",
     "2M reads/mo cap",
     "Backstop only",
     5000, 25, 25,
     "Buffer / backstop. Grok Live Search handles 80% of X needs; this covers the rest.",
     "Drop entirely — but redundancy is cheap insurance.",
     "https://docs.x.com/x-api/getting-started/about-x-api"),

    # ---------- L4 — Enrichment & entity ----------
    (19, "L4", "Diffbot",
     "Knowledge Graph + Article + NLP. Primary company enrichment. ~3 credits per candidate.",
     "Startup (v1) → Plus (P2)",
     299, "$0.0012 / credit",
     "$1.20", "$12", "$120",
     "5 calls/sec (Startup), 25 calls/sec (Plus)",
     "~18K credits (v1), ~75K credits (P2)",
     18000, 299, 899,
     "WINNER. Cleanest structured-data extraction on the market. Phase 2 needs the $899 Plus plan for live NLS lookups.",
     "Clearbit ($999/mo+) — comparable but limited to companies, no article extract.",
     "https://www.diffbot.com/pricing/"),

    (20, "L4", "Splink (entity resolution)",
     "Probabilistic record linkage / dedup. Open source, runs on our Fly.io machine.",
     "Open source",
     0, "Compute only (counted in Fly.io)",
     "n/a", "n/a", "n/a",
     "Bounded by compute",
     "All 6K candidates/mo",
     6000, 0, 0,
     "WINNER. Industry-standard probabilistic ER (UK gov-built). No vendor cost.",
     "Senzing (paid commercial ER) — $50K+/yr enterprise, overkill at our scale.",
     "https://moj-analytical-services.github.io/splink/"),

    # ---------- L5 — LLM inference ----------
    (21, "L5", "OpenAI — GPT-4o-mini",
     "Signal parsing — extract entities from raw text. ~6K calls/mo.",
     "Pay-per-token",
     0, "$0.15 in / $0.60 out per 1M tokens",
     "n/a", "n/a", "n/a",
     "Tier-based: Tier 1 = 500 RPM, Tier 5 = 30K RPM",
     "~3K input + 500 output tokens / candidate",
     6000, 5, 5,
     "WINNER for parsing. Cheapest reliable structured-output model.",
     "Gemini 2.5 Flash-Lite ($0.10/$0.40) — cheaper but weaker JSON adherence in our tests.",
     "https://openai.com/api/pricing/"),

    (22, "L5", "OpenAI — text-embedding-3-small",
     "Embeddings for semantic search + dedup.",
     "Pay-per-token",
     0, "$0.02 / 1M tokens",
     "n/a", "n/a", "n/a",
     "Tier-based, very generous",
     "~5K tokens / candidate",
     6000, 1, 1,
     "WINNER. Industry standard, integrates cleanly with pgvector. Negligible cost.",
     "Cohere embed-v3 — comparable quality, similar pricing.",
     "https://openai.com/api/pricing/"),

    (23, "L5", "Anthropic — Claude Sonnet 4.6",
     "Score justification + outreach drafting. Best-in-class BD prose.",
     "Pay-per-token",
     0, "$3 in / $15 out per 1M tokens",
     "n/a", "n/a", "n/a",
     "Tier-based, generous",
     "~8K in + 1K out / justified candidate",
     1800, 71, 71,
     "WINNER for prose-quality steps. Cleaner, more diplomatic output than GPT-4 — ships to executives without rewrites.",
     "GPT-4o ($2.50/$10) — slightly cheaper, prose less polished. Worth the marginal cost.",
     "https://www.anthropic.com/pricing#api"),

    (24, "L5", "Perplexity Sonar Pro / Deep Research",
     "Phase 2 only — natural-language search frontend with built-in real-time web grounding.",
     "Pay-per-token",
     0, "Sonar Pro: $3 in / $15 out · Deep Research: $2/$8 + reasoning surcharge",
     "n/a", "n/a", "n/a",
     "Tier 1 = 50 RPM, scales with spend",
     "~2,200 queries × $0.05 avg",
     2200, 0, 110,
     "WINNER for Phase 2 NLS. Built-in real-time web grounding + citations — saves rebuilding search-and-cite.",
     "Build it ourselves on EXA + Claude — possible but adds 2–3 weeks of dev.",
     "https://docs.perplexity.ai/docs/getting-started/pricing"),

    # ---------- INFRA ----------
    (25, "INFRA", "Supabase — Team plan",
     "Postgres + pgvector + pgvectorscale + auth + storage. SOC2, 28-day PITR, audit logs.",
     "Team",
     599, "$599 base + usage",
     "n/a", "n/a", "n/a",
     "Per-project compute scales",
     "Full primary DB + vectors",
     0, 599, 799,
     "WINNER for premium / enterprise-ready. One DB, one bill, SOC2 + multi-tenant via RLS.",
     "Pro ($25) — saves $574/mo but no SOC2/HIPAA, no 28-day PITR, no audit logs.",
     "https://supabase.com/pricing"),

    (26, "INFRA", "Upstash Redis + Prod Pack",
     "Job queue (BullMQ), rate limiters, dedup cache, session.",
     "Pay-as-you-go + Prod Pack",
     200, "$0.20/100K cmd · $0.03/GB · +$200 Prod Pack",
     "n/a", "n/a", "n/a",
     "Per-DB throughput limits",
     "~150M cmd/mo + 100GB bw",
     150000000, 230, 230,
     "WINNER. Prod Pack adds SOC2/HIPAA/encryption-at-rest. Serverless Redis pricing is unbeatable.",
     "Redis Cloud Essentials ($7+/mo) — cheaper but no SOC2 in lower tiers.",
     "https://upstash.com/pricing"),

    (27, "INFRA", "Cloudflare R2",
     "Object storage for raw signal blobs, scraped HTML, generated PDFs.",
     "Pay-as-you-go",
     0, "$0.015/GB-mo storage · FREE egress",
     "n/a", "n/a", "n/a",
     "Operations: $4.50/M Class A, $0.36/M Class B",
     "~50 GB storage + light writes",
     0, 5, 10,
     "WINNER. Free egress vs S3 ~$0.09/GB is a game-changer. SOC2/ISO27001 included.",
     "AWS S3 — adds significant egress cost; only choose for tight AWS-shop integration.",
     "https://developers.cloudflare.com/r2/pricing/"),

    (28, "INFRA", "Inngest",
     "Workflow orchestration — durable functions, retries, fan-out, scheduling.",
     "Paid (~Base $75)",
     75, "$75 base, 50K free executions, then per-exec",
     "n/a", "n/a", "n/a",
     "Per-plan execution caps",
     "~200K executions/mo",
     200000, 75, 150,
     "WINNER. Cleanest TypeScript-native workflow engine. Phase 2 needs higher tier for NLS pipelines.",
     "Temporal ($100+/mo) — more powerful but heavier ops burden.",
     "https://www.inngest.com/pricing"),

    (29, "INFRA", "Meilisearch Cloud — Pro",
     "Full-text + hybrid search across companies + signals for the dashboard.",
     "Pro",
     300, "$300 base + overages",
     "n/a", "n/a", "n/a",
     "Per-instance throughput",
     "All dashboard search",
     0, 300, 300,
     "WINNER. Hybrid search (BM25 + vectors) under one engine. Faster than Algolia, cheaper at scale.",
     "Algolia ($0.50/1K records + $1.50/1K ops) — premium but 3–5× cost at our document count.",
     "https://www.meilisearch.com/pricing"),

    (30, "INFRA", "Langfuse — Pro",
     "LLM observability — every prompt, response, latency, cost tracked. Critical for tuning.",
     "Pro",
     199, "$199 base + $8/100K units overage",
     "n/a", "n/a", "n/a",
     "Throughput per plan",
     "~500K observations/mo",
     500000, 199, 199,
     "WINNER. MIT-licensed, ClickHouse-backed (acquired Jan 2026). Best LLM tracing for cost + quality eval.",
     "Helicone ($25+/mo) — cheaper, less feature depth on evals + datasets.",
     "https://langfuse.com/pricing"),

    (31, "INFRA", "Fly.io",
     "Backend workers (4 shared-CPU machines + 1 dedicated CPU for Splink).",
     "Pay-as-you-go",
     0, "From $0.0027/hr shared CPU; $99/mo perf machines",
     "n/a", "n/a", "n/a",
     "Per-machine resources",
     "~4 always-on + 1 dedicated",
     0, 220, 320,
     "WINNER. Per-second billing, scale-to-zero, global edge. Phase 2 adds NLS query workers.",
     "Railway ($5+/mo + usage) — simpler but worse multi-region story.",
     "https://fly.io/docs/about/pricing/"),

    (32, "INFRA", "Vercel — Pro",
     "Frontend dashboard hosting, edge functions, ISR. ~5 seats.",
     "Pro",
     130, "$20/seat + usage; $0.15/GB bw overage",
     "n/a", "n/a", "n/a",
     "1TB bw + 10M edge req included",
     "Dashboard frontend + edge fns",
     0, 130, 200,
     "WINNER for Next.js shop. Phase 2 needs more function CPU for NLS streaming.",
     "Cloudflare Pages — cheaper but weaker Next.js ecosystem support.",
     "https://vercel.com/pricing"),

    (33, "INFRA", "Sentry — Team",
     "Error tracking + performance on frontend + backend.",
     "Team",
     26, "$26 base + per-event overage",
     "n/a", "n/a", "n/a",
     "Plan event quotas",
     "~50K errors/mo",
     50000, 26, 26,
     "WINNER. Industry standard for cross-stack error tracking with great Next.js + Python SDKs.",
     "Bugsnag, Rollbar — comparable, no compelling reason to switch.",
     "https://sentry.io/pricing/"),
]

# Write rows
r = 5
for entry in catalog:
    fill_ = FILL_ALT if r % 2 else None
    write_row(ws, r, list(entry), font=F_BODY, fill_=fill_, align=A_TOP_L, border=box(),
              fmts=[None, None, None, None, None,
                    '"$"#,##0', None, None, None, None,
                    None, None, "#,##0",
                    '"$"#,##0', '"$"#,##0',
                    None, None, None])
    # Make verdict cell bold
    ws.cell(row=r, column=16).font = F_BODY_B
    # Source link
    src_cell = ws.cell(row=r, column=18)
    if entry[17] and entry[17].startswith("http"):
        src_cell.hyperlink = entry[17]
        src_cell.font = F_LINK
    ws.row_dimensions[r].height = 60
    r += 1

# Subtotals
sub_external_v1 = sum(e[13] for e in catalog if e[1] in ("L1","L2","L3","L4"))
sub_external_p2 = sum(e[14] for e in catalog if e[1] in ("L1","L2","L3","L4"))
sub_llm_v1      = sum(e[13] for e in catalog if e[1] == "L5")
sub_llm_p2      = sum(e[14] for e in catalog if e[1] == "L5")
sub_infra_v1    = sum(e[13] for e in catalog if e[1] == "INFRA")
sub_infra_p2    = sum(e[14] for e in catalog if e[1] == "INFRA")
total_v1        = sub_external_v1 + sub_llm_v1 + sub_infra_v1
total_p2        = sub_external_p2 + sub_llm_p2 + sub_infra_p2

# Subtotal block
def subtotal_row(ws, r, label, v1, p2, fill_):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
    c = ws.cell(row=r, column=1, value=label)
    c.font = F_SUBHEAD
    c.fill = fill_
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.border = box()
    for col in range(2, 14):
        ws.cell(row=r, column=col).fill = fill_
        ws.cell(row=r, column=col).border = box()
    cv1 = ws.cell(row=r, column=14, value=v1); cv1.font = F_SUBHEAD; cv1.fill = fill_; cv1.number_format = '"$"#,##0'; cv1.alignment = A_RIGHT; cv1.border = box()
    cp2 = ws.cell(row=r, column=15, value=p2); cp2.font = F_SUBHEAD; cp2.fill = fill_; cp2.number_format = '"$"#,##0'; cp2.alignment = A_RIGHT; cp2.border = box()
    for col in range(16, 19):
        ws.cell(row=r, column=col).fill = fill_
        ws.cell(row=r, column=col).border = box()
    ws.row_dimensions[r].height = 22

subtotal_row(ws, r,   "External Data APIs (L1+L2+L3+L4) subtotal", sub_external_v1, sub_external_p2, FILL_SUBHEAD)
subtotal_row(ws, r+1, "LLM Inference (L5) subtotal",                sub_llm_v1,      sub_llm_p2,      FILL_SUBHEAD)
subtotal_row(ws, r+2, "Core Infrastructure subtotal",               sub_infra_v1,    sub_infra_p2,    FILL_SUBHEAD)

# Grand total
gr = r + 3
ws.merge_cells(start_row=gr, start_column=1, end_row=gr, end_column=13)
c = ws.cell(row=gr, column=1, value="GRAND TOTAL — monthly per deployment")
c.font = F_TOTAL; c.fill = FILL_TOTAL; c.alignment = Alignment(horizontal="right", vertical="center", indent=1); c.border = box()
for col in range(2, 14):
    ws.cell(row=gr, column=col).fill = FILL_TOTAL
    ws.cell(row=gr, column=col).border = box()
cv1 = ws.cell(row=gr, column=14, value=total_v1); cv1.font = F_TOTAL; cv1.fill = FILL_TOTAL; cv1.number_format = '"$"#,##0'; cv1.alignment = A_RIGHT; cv1.border = box()
cp2 = ws.cell(row=gr, column=15, value=total_p2); cp2.font = F_TOTAL; cp2.fill = FILL_TOTAL; cp2.number_format = '"$"#,##0'; cp2.alignment = A_RIGHT; cp2.border = box()
for col in range(16, 19):
    ws.cell(row=gr, column=col).fill = FILL_TOTAL
    ws.cell(row=gr, column=col).border = box()
ws.row_dimensions[gr].height = 28

# Freeze
ws.freeze_panes = "C5"
ws.sheet_view.zoomScale = 95


# ============================================================
# TAB 3 — PER-CANDIDATE COST FLOW
# ============================================================
ws = wb.create_sheet("Per-Candidate Flow")
ws.sheet_properties.tabColor = ACCENT_GREEN
ws.sheet_view.showGridLines = False

set_col_widths(ws, [4, 36, 38, 18, 18, 28])

title_block(
    ws,
    "Per-Candidate Cost Flow — One Company Through the Pipeline",
    "Tracking exactly what fires when one candidate enters NORAD. Volumes amortized at 200 candidates/day.",
    last_col_letter="F",
)

header_row(ws, 4, ["Stage", "What happens", "API calls / tools fired", "Cost / candidate", "Confidence", "Notes"], height=36)

flow = [
    (1, "Signal ingested",
     "One signal among many lands in the queue (filing / news / patent / social post)",
     "1 SerpApi search + ~0.5 Firecrawl scrapes",
     0.010, "High",
     "Amortized — many signals per ingestion run hit zero or near-zero."),
    (2, "Entity extraction & dedup",
     "Raw text parsed → company entity → embedded → matched against existing companies",
     "1 GPT-4o-mini parse + 1 OpenAI embedding + 1 Splink ER pass",
     0.001, "High",
     "GPT-4o-mini at $0.0008/parse + embedding $0.0001."),
    (3, "Enrichment",
     "Full company profile assembled — Knowledge Graph, semantic web search, content extraction, structured pulls",
     "3 Diffbot credits + 5 EXA searches + 10 EXA contents + 1 Firecrawl extract",
     0.050, "High",
     "Largest line item per candidate. EXA is the biggest contributor."),
    (4, "Closed-platform pull",
     "LinkedIn profiles for company + founders, optional social/PitchBook",
     "3 LinkedIn + 0.3 PitchBook + 0.2 social profile",
     0.014, "Medium",
     "Average — some candidates need none, some need all."),
    (5, "Hard filter",
     "Deterministic rules drop ~70%: vertical, geo, size, blocked tags",
     "No API calls",
     0.000, "High",
     "Pure compute. ~30% of candidates pass through to step 6."),
    (6, "Justification (~30% of candidates)",
     "Claude Sonnet 4.6 writes scoring justification + 'why this matters'",
     "1 Claude Sonnet 4.6 call (~8K in + 1K out)",
     0.012, "High",
     "Cost = $0.039/justification × 30% pass rate = $0.012 amortized."),
    (7, "Storage & indexing",
     "Persist to Supabase (rows + vectors), Meilisearch index update, R2 raw blob",
     "DB writes + vector upsert + search index + object storage",
     0.005, "Medium",
     "Amortized fixed costs across 6,000 candidates/mo."),
]

r = 5
for stage, what, action, calls, cost, conf, notes in flow:
    fill_ = FILL_ALT if r % 2 else None
    write_row(ws, r, [stage, what, calls, cost, conf, notes],
              font=F_BODY, fill_=fill_, align=A_TOP_L, border=box(),
              fmts=[None, None, None, '"$"#,##0.0000', None, None])
    # Confidence color
    cc = ws.cell(row=r, column=5)
    cc.font = F_BODY_B
    if conf == "High":
        cc.fill = FILL_GOOD
    elif conf == "Medium":
        cc.fill = FILL_WARN
    else:
        cc.fill = FILL_BAD
    ws.row_dimensions[r].height = 50
    r += 1

# Total row
total_per_candidate = sum(f[4] for f in flow)
write_row(ws, r,
          ["", "TOTAL PER CANDIDATE PROCESSED", "", total_per_candidate, "", ""],
          font=F_TOTAL, fill_=FILL_TOTAL, align=A_RIGHT, border=box(),
          fmts=[None, None, None, '"$"#,##0.0000', None, None])
ws.row_dimensions[r].height = 26

# Translation block
r += 2
section_label(ws, r, "What that translates to", "F")
r += 1
ws.row_dimensions[r].height = 8
r += 1
header_row(ws, r, ["Metric", "Calculation", "", "Value", "", "Plain-English meaning"], height=28)
trans = [
    ("Per candidate processed", f"${total_per_candidate:.4f} per candidate",
     "", total_per_candidate, "",
     "Cost the engine incurs to evaluate one company end-to-end."),
    ("Per qualified opportunity (≥70 score)", "$3,042 / 80 qualified per month",
     "", 38.0, "",
     "Cost per opportunity that actually reaches the operator review queue."),
    ("Per Elite alert (≥85 score)", "$3,042 / 12 Elite per month",
     "", 254.0, "",
     "Cost per top-of-funnel opportunity surfaced to client deal teams."),
    ("Per BD-team-hour saved", "(see Assumptions tab) ~6 hr/Elite",
     "", 42.0, "",
     "Approx engine cost per hour of analyst research replaced."),
]
r += 1
for m, calc, _, val, _, plain in trans:
    fill_ = FILL_ALT if r % 2 else None
    write_row(ws, r, [m, calc, "", val, "", plain],
              font=F_BODY_B, fill_=fill_, align=A_LEFT, border=box(),
              fmts=[None, None, None, '"$"#,##0.00', None, None])
    ws.row_dimensions[r].height = 22
    r += 1

ws.freeze_panes = "A5"


# ============================================================
# TAB 4 — WORKFLOW SCENARIOS
# ============================================================
ws = wb.create_sheet("Workflow Scenarios")
ws.sheet_properties.tabColor = ACCENT_GREEN
ws.sheet_view.showGridLines = False

set_col_widths(ws, [4, 36, 32, 16, 16, 28])

title_block(
    ws,
    "Workflow Scenarios — Cost of Chained API Calls",
    "Three end-to-end workflows showing what stacks of APIs cost when chained in sequence.",
    last_col_letter="F",
)

# ---------- Scenario A: Standard pipeline (already covered in Per-Candidate) ----------
section_label(ws, 4, "Scenario A — Standard discovery pipeline (one candidate end-to-end)", "F")
ws.row_dimensions[5].height = 6
header_row(ws, 6, ["Step", "Tool chained", "Action", "Calls", "Cost", "Notes"], height=30)
scenario_a = [
    (1, "SerpApi → Firecrawl", "Find new signal → fetch full page", "1 + 0.5", 0.0104, "Per signal ingested"),
    (2, "GPT-4o-mini → embedding-3-small", "Parse entity → embed for dedup", "1 + 1", 0.0009, "Always fires"),
    (3, "Diffbot → EXA → Firecrawl", "Build full company profile", "3 + 15 + 1", 0.0500, "All ~3 of these chained"),
    (4, "Apify LinkedIn → Apify PitchBook", "Add founder + funding context", "3 + 0.3", 0.0140, "PitchBook fires only on private targets"),
    (5, "Claude Sonnet 4.6", "Write scoring justification", "0.3 calls", 0.0117, "Only for ~30% that pass hard filter"),
    (6, "Supabase + Meilisearch + R2", "Persist + index + archive", "writes only", 0.0050, "Amortized infra"),
]
r = 7
sub_a = 0
for step, chain, act, calls, cost, notes in scenario_a:
    fill_ = FILL_ALT if r % 2 else None
    write_row(ws, r, [step, chain, act, calls, cost, notes],
              font=F_BODY, fill_=fill_, align=A_TOP_L, border=box(),
              fmts=[None, None, None, None, '"$"#,##0.0000', None])
    ws.row_dimensions[r].height = 32
    sub_a += cost
    r += 1
write_row(ws, r,
          ["", "SCENARIO A TOTAL — per candidate", "", "", sub_a, "Standard end-to-end pipeline"],
          font=F_TOTAL, fill_=FILL_TOTAL, align=A_RIGHT, border=box(),
          fmts=[None, None, None, None, '"$"#,##0.0000', None])
ws.row_dimensions[r].height = 26
r += 2

# ---------- Scenario B: Deep Research workflow (analyst-triggered) ----------
section_label(ws, r, "Scenario B — Deep Research workflow (analyst hits 'Investigate' on a candidate)", "F")
r += 1; ws.row_dimensions[r].height = 6; r += 1
header_row(ws, r, ["Step", "Tool chained", "Action", "Calls", "Cost", "Notes"], height=30)
r += 1
scenario_b = [
    (1, "Perplexity Sonar Deep Research", "Multi-step web research with citations", "1 deep query", 0.180, "~15K in + 8K out + reasoning surcharge"),
    (2, "EXA Deep Search", "Targeted semantic search across niche sources", "5 queries", 0.075, "$15/1K Deep Search tier"),
    (3, "EXA Contents", "Pull full text of top 30 hits", "30 contents", 0.030, "Feeds the LLM"),
    (4, "Diffbot KG enhance", "Pull every related entity (people, products, competitors)", "20 credits", 0.024, "Wider net than standard enrichment"),
    (5, "Apify LinkedIn (full team)", "Scrape full leadership + recent hires", "30 profiles", 0.090, "Per-result actor"),
    (6, "Apify PitchBook (full)", "Funding history + cap table + competitors", "1 deep pull", 0.045, "Multi-page actor run"),
    (7, "Grok Live Search", "Founder + brand X mentions, last 90 days", "1 query (~5K tok)", 0.014, "Native X access"),
    (8, "Claude Sonnet 4.6", "Synthesize 6-section investment-memo-style output", "1 call (~30K in + 4K out)", 0.150, "Long context, full report"),
    (9, "Supabase + R2 + PDF", "Persist memo + render PDF", "writes only", 0.005, "Amortized"),
]
sub_b = 0
for step, chain, act, calls, cost, notes in scenario_b:
    fill_ = FILL_ALT if r % 2 else None
    write_row(ws, r, [step, chain, act, calls, cost, notes],
              font=F_BODY, fill_=fill_, align=A_TOP_L, border=box(),
              fmts=[None, None, None, None, '"$"#,##0.0000', None])
    ws.row_dimensions[r].height = 32
    sub_b += cost
    r += 1
write_row(ws, r,
          ["", "SCENARIO B TOTAL — per Deep Research run", "", "", sub_b,
           "Triggered manually by analyst. Compare to $500–2K human equivalent."],
          font=F_TOTAL, fill_=FILL_TOTAL, align=A_RIGHT, border=box(),
          fmts=[None, None, None, None, '"$"#,##0.00', None])
ws.row_dimensions[r].height = 26
r += 2

# ---------- Scenario C: Phase 2 NLS query ----------
section_label(ws, r, "Scenario C — Phase 2 natural-language query (analyst types a question)", "F")
r += 1; ws.row_dimensions[r].height = 6; r += 1
header_row(ws, r, ["Step", "Tool chained", "Action", "Calls", "Cost", "Notes"], height=30)
r += 1
scenario_c = [
    (1, "Meilisearch hybrid",         "Internal search across NORAD's existing data", "1 query", 0.0001, "Primary path — answer often lives in our DB"),
    (2, "pgvector ANN",               "Semantic match on candidate vectors",          "1 query", 0.0001, "Free at our scale, runs on Supabase"),
    (3, "Perplexity Sonar Pro",       "Live-web grounding when DB miss",              "1 query (~2K in + 1K out)", 0.0210, "$3 in / $15 out per 1M tokens"),
    (4, "EXA Search",                 "Fresh sources beyond DB",                      "3 queries", 0.0210, "Phase 2 NLS layer"),
    (5, "Grok Live Search",           "X-specific question grounding",                "0.3 queries (~3K tok)", 0.0030, "Only when query mentions X / social"),
    (6, "Claude Sonnet 4.6",          "Final answer synthesis with citations",        "1 call (~6K in + 1K out)", 0.0330, "Quality answer + citations"),
    (7, "Langfuse trace",             "Log the whole chain for tuning",               "1 trace + ~6 obs", 0.0001, "Counted in plan units"),
]
sub_c = 0
for step, chain, act, calls, cost, notes in scenario_c:
    fill_ = FILL_ALT if r % 2 else None
    write_row(ws, r, [step, chain, act, calls, cost, notes],
              font=F_BODY, fill_=fill_, align=A_TOP_L, border=box(),
              fmts=[None, None, None, None, '"$"#,##0.0000', None])
    ws.row_dimensions[r].height = 32
    sub_c += cost
    r += 1
write_row(ws, r,
          ["", "SCENARIO C TOTAL — per natural-language query", "", "", sub_c,
           "~$0.05/query at 2,200 queries/mo = $110/mo Phase 2 line item"],
          font=F_TOTAL, fill_=FILL_TOTAL, align=A_RIGHT, border=box(),
          fmts=[None, None, None, None, '"$"#,##0.0000', None])
ws.row_dimensions[r].height = 26

# Summary comparison
r += 2
section_label(ws, r, "Workflow comparison at a glance", "F")
r += 1; ws.row_dimensions[r].height = 6; r += 1
header_row(ws, r, ["Workflow", "Trigger", "Avg cost", "Monthly volume", "Monthly $", "Comparable manual cost"], height=30)
r += 1
comp = [
    ("A — Standard pipeline",     "Automatic, every candidate",        0.092, 6000, "$552", "$50–200 / company manually"),
    ("B — Deep Research run",     "Analyst-triggered, on-demand",      0.613, 50,   "$31",  "$500–2,000 / report manually"),
    ("C — NLS query (Phase 2)",   "Analyst typing in dashboard",       0.078, 2200, "$172", "$5–20 / query (analyst time)"),
]
for w, t, c_, v, m, manual in comp:
    fill_ = FILL_ALT if r % 2 else None
    write_row(ws, r, [w, t, c_, v, m, manual],
              font=F_BODY_B, fill_=fill_, align=A_LEFT, border=box(),
              fmts=[None, None, '"$"#,##0.000', "#,##0", None, None])
    ws.row_dimensions[r].height = 22
    r += 1

ws.freeze_panes = "A4"


# ============================================================
# TAB 5 — VOLUME SENSITIVITY
# ============================================================
ws = wb.create_sheet("Volume Sensitivity")
ws.sheet_properties.tabColor = ACCENT_AMBR
ws.sheet_view.showGridLines = False

set_col_widths(ws, [4, 26, 16, 16, 16, 16, 16, 16, 28])

title_block(
    ws,
    "Volume Sensitivity — How Cost Moves with Throughput",
    "Per-deployment monthly cost at five different daily candidate volumes. Shows fixed vs variable split.",
    last_col_letter="I",
)

header_row(ws, 4, ["#", "Bucket", "50/day", "100/day", "200/day (REC)", "500/day", "1,000/day", "2,000/day", "Notes"], height=32)

# Rough scaling logic
# - External APIs scale ~linearly with volume (per-candidate cost driver)
# - LLM scales linearly
# - Infra scales sub-linearly: ~30% per 2× volume
def scale_external(base_v1_at_200, daily):
    return round(base_v1_at_200 * (daily / 200))

def scale_llm(base_v1_at_200, daily):
    return round(base_v1_at_200 * (daily / 200))

def scale_infra(base_v1_at_200, daily):
    # Sub-linear: floor at 0.7 × base; ceiling at +60% per 5× volume
    multiplier = 1 + 0.30 * ((daily / 200) - 1)
    multiplier = max(0.85, multiplier)
    return round(base_v1_at_200 * multiplier)

volumes = [50, 100, 200, 500, 1000, 2000]
v1_external_base = sub_external_v1
v1_llm_base = sub_llm_v1
v1_infra_base = sub_infra_v1

ext_row = ["External Data APIs"] + [scale_external(v1_external_base, d) for d in volumes] + ["Scales linearly per candidate"]
llm_row = ["LLM Inference"]      + [scale_llm(v1_llm_base, d) for d in volumes]            + ["Scales linearly per candidate"]
inf_row = ["Core Infrastructure"]+ [scale_infra(v1_infra_base, d) for d in volumes]        + ["Sub-linear: ~30% per 2× volume"]
tot_row = ["TOTAL MONTHLY"]      + [ext_row[i] + llm_row[i] + inf_row[i] for i in range(1, 7)] + ["Per-deployment monthly bill"]

per_cand = ["Per candidate processed"] + [round(tot_row[i] / (volumes[i-1] * 30), 4) for i in range(1, 7)] + ["v1 cost / candidate"]
per_qual = ["Per qualified opportunity"] + [round(tot_row[i] / max(1, round(volumes[i-1] * 30 * 0.0133)), 2) for i in range(1, 7)] + ["≥70 score, ~1.33% of processed"]
per_eli  = ["Per Elite alert"]           + [round(tot_row[i] / max(1, round(volumes[i-1] * 30 * 0.002)), 2) for i in range(1, 7)]  + ["≥85 score, ~0.2% of processed"]

rows = [ext_row, llm_row, inf_row, tot_row, per_cand, per_qual, per_eli]
labels_idx_total = 3  # "TOTAL MONTHLY"

r = 5
for i, row_data in enumerate(rows):
    is_total = (i == labels_idx_total)
    fill_ = FILL_TOTAL if is_total else (FILL_ALT if r % 2 else None)
    font_ = F_TOTAL if is_total else F_BODY
    is_per = i > labels_idx_total
    money_fmt = '"$"#,##0' if not is_per else '"$"#,##0.0000'
    if i == 5: money_fmt = '"$"#,##0.00'
    if i == 6: money_fmt = '"$"#,##0.00'
    fmts = [None, None] + [money_fmt]*6 + [None]
    # Insert # column
    full_row = [i+1] + row_data
    write_row(ws, r, full_row, font=font_, fill_=fill_, align=A_LEFT, border=box(), fmts=fmts)
    ws.row_dimensions[r].height = 24
    # Highlight 200/day column
    rec_cell = ws.cell(row=r, column=5)
    if not is_total:
        rec_cell.fill = FILL_GOOD
    rec_cell.font = Font(name="Calibri", size=10, bold=True,
                         color=WHITE if is_total else "222222")
    r += 1

# Add a note row
r += 1
ws.merge_cells(f"A{r}:I{r}")
c = ws.cell(row=r, column=1,
            value="Highlighted column (200/day) is the recommended baseline. "
                  "Numbers are v1 stack only; Phase 2 adds ~$1,300/mo across all volumes.")
c.font = F_NOTE
c.alignment = A_LEFT

ws.freeze_panes = "C5"


# ============================================================
# TAB 6 — LLM MODELS REFERENCE
# ============================================================
ws = wb.create_sheet("LLM Models Reference")
ws.sheet_properties.tabColor = ACCENT_AMBR
ws.sheet_view.showGridLines = False

set_col_widths(ws, [4, 30, 14, 14, 14, 14, 38, 22])

title_block(
    ws,
    "LLM Models Reference — Every Frontier Model Considered (April–May 2026)",
    "All pricing per 1M tokens. Used to justify our model picks per pipeline step.",
    last_col_letter="H",
)

header_row(ws, 4, ["#", "Model", "Input $/1M", "Output $/1M", "Context", "Used in NORAD?", "Best for", "Verdict in our stack"], height=32)

models = [
    ("OpenAI GPT-4o-mini",      0.15,  0.60,  "128K", "YES",  "Cheap structured parsing — our parser",                "PICKED for L5 parsing"),
    ("OpenAI GPT-4o",           2.50, 10.00,  "128K", "no",   "Legacy generalist",                                    "Superseded by Sonnet for prose"),
    ("OpenAI GPT-4.1",          2.00,  8.00,  "1M",   "no",   "General reasoning alt",                                "No quality edge over Sonnet at 3× cost"),
    ("OpenAI text-embedding-3-small", 0.02, 0.0, "n/a", "YES", "Embeddings for pgvector",                              "PICKED for L5 embeddings"),
    ("Anthropic Claude Haiku 4.5", 1.00, 5.00, "200K", "no",  "Fast cheap reasoning",                                 "Sonnet wins on quality at 3× cost"),
    ("Anthropic Claude Sonnet 4.6", 3.00, 15.00, "1M",  "YES", "Justification + outreach prose",                       "PICKED for L5 justification + drafts"),
    ("Anthropic Claude Opus 4.7",   5.00, 25.00, "1M",  "no",  "Top-end reasoning, expensive",                          "Reserved — too costly for our volume"),
    ("Google Gemini 2.5 Flash-Lite", 0.10, 0.40, "1M", "no",  "Cheapest tier alt to GPT-4o-mini",                      "Weaker JSON adherence in tests"),
    ("Google Gemini 2.5 Flash",      0.30, 2.50, "1M", "no",  "Mid-tier alt",                                          "No compelling edge"),
    ("Google Gemini 2.5 Pro (≤200K)", 1.25, 10.00, "1M","no", "Heavy alt to Sonnet",                                   "Sonnet has better long-context recall"),
    ("xAI Grok 4.3",                 1.25,  2.50, "1M", "YES","X + live web grounding (Live Search built in)",         "PICKED for L3 X access"),
    ("Perplexity Sonar",             1.00,  1.00, "n/a","no", "Lightweight live-web Q&A",                              "Sonar Pro better for our NLS depth"),
    ("Perplexity Sonar Pro",         3.00, 15.00, "n/a","YES (P2)","Phase 2 NLS query frontend",                       "PICKED for Phase 2 NLS"),
    ("Perplexity Sonar Reasoning Pro", 2.00, 8.00, "n/a","no", "Reasoning over live web",                              "Deep Research is more applicable"),
    ("Perplexity Sonar Deep Research", 2.00, 8.00, "n/a","YES (B)","Multi-step deep research workflow",                "PICKED for Deep Research scenario"),
]
r = 5
for i, m in enumerate(models, start=1):
    fill_ = FILL_ALT if r % 2 else None
    write_row(ws, r, [i] + list(m),
              font=F_BODY, fill_=fill_, align=A_LEFT, border=box(),
              fmts=[None, None, '"$"#,##0.00', '"$"#,##0.00', None, None, None, None])
    # Color "Used in NORAD?" cell
    uc = ws.cell(row=r, column=6)
    if "YES" in str(m[4]):
        uc.fill = FILL_GOOD; uc.font = F_BODY_B
    else:
        uc.font = F_NOTE
    # Bold verdict
    ws.cell(row=r, column=8).font = F_BODY_B
    ws.row_dimensions[r].height = 26
    r += 1

ws.freeze_panes = "A5"


# ============================================================
# TAB 7 — DATABASE COMPARISON
# ============================================================
ws = wb.create_sheet("DB & Vector Comparison")
ws.sheet_properties.tabColor = ACCENT_AMBR
ws.sheet_view.showGridLines = False

set_col_widths(ws, [4, 30, 18, 16, 14, 18, 38, 20])

title_block(
    ws,
    "Database & Vector Storage Comparison",
    "Side-by-side answer to 'Supabase or something else?' — premium pick + alternatives evaluated.",
    last_col_letter="H",
)

header_row(ws, 4, ["#", "Option", "Monthly cost (v1)", "Vector quality", "Multi-tenant?", "SOC2 / HIPAA", "Notes", "Verdict"], height=32)

opts = [
    ("Supabase Pro + pgvector + pgvectorscale", 35, "Production-grade", "Yes (RLS)", "No",
     "Best-bang-for-buck. Saves $574/mo vs Team but no compliance certs.",
     "Strong v1 if compliance not required"),
    ("Supabase Team + pgvector + pgvectorscale", 599, "Production-grade", "Yes (RLS)", "Yes (SOC2 + HIPAA)",
     "One product, one bill. SOC2 + 28-day PITR + audit logs + SSO.",
     "PICKED — premium / enterprise-ready"),
    ("Neon Postgres + Pinecone", 120, "Excellent (dedicated)", "Manual", "Add-on",
     "Splits stack into two products. More ops, more bills, harder joins between vectors and rows.",
     "Adds complexity, no quality win"),
    ("Pinecone Standard alone", 50, "Best dedicated", "Yes (namespaces)", "Yes",
     "Need separate Postgres anyway → adds bill. Makes sense at 100M+ vectors only.",
     "Wrong tool at our scale"),
    ("Self-hosted Postgres + pgvector on Fly.io", 120, "Production-grade", "Manual", "DIY",
     "Saves ~$80/mo over Pro but adds DBA burden + no managed backups.",
     "Not worth the ops cost"),
    ("PlanetScale (Vitess) + Pinecone", 200, "Excellent (Pinecone)", "Yes", "Yes",
     "Great horizontal scale but no Postgres extensions (no pgvectorscale, no PostGIS, no Splink-friendly joins).",
     "Loses our extension story"),
    ("Snowflake / Databricks + vector", 1500, "Good but expensive", "Yes", "Yes",
     "Massive overkill at our scale. Compute-per-query model bleeds money for OLTP workloads.",
     "Wrong category — analytics warehouse"),
]
r = 5
for i, o in enumerate(opts, start=1):
    fill_ = FILL_ALT if r % 2 else None
    is_picked = "PICKED" in o[6]
    write_row(ws, r, [i] + list(o),
              font=F_BODY, fill_=fill_, align=A_TOP_L, border=box(),
              fmts=[None, None, '"$"#,##0', None, None, None, None, None])
    if is_picked:
        for col in range(1, 9):
            ws.cell(row=r, column=col).fill = FILL_GOOD
            ws.cell(row=r, column=col).font = F_BODY_B
    ws.cell(row=r, column=8).font = F_BODY_B
    ws.row_dimensions[r].height = 50
    r += 1

# Why pgvector reasoning block
r += 2
section_label(ws, r, "Why pgvector over a dedicated vector DB", "H")
r += 1; ws.row_dimensions[r].height = 6; r += 1
reasons = [
    "1. One database, one query language, one transaction boundary.",
    "2. Splink (entity resolution) needs joins between vectors and structured data — trivially easy in Postgres, painful across two products.",
    "3. pgvectorscale (built by Timescale, free extension) gives Pinecone-equivalent recall at our scale (~6K candidates/mo, ~500K vectors steady-state per deployment).",
    "4. Pinecone makes sense at 100M+ vectors — we are nowhere near that ceiling.",
    "5. RLS in Supabase gives us multi-tenant isolation for free in Phase 2 — Pinecone uses namespaces, fine but adds layer.",
]
for line in reasons:
    ws.merge_cells(f"A{r}:H{r}")
    c = ws.cell(row=r, column=1, value=line)
    c.font = F_BODY
    c.alignment = A_LEFT
    ws.row_dimensions[r].height = 20
    r += 1

ws.freeze_panes = "A5"


# ============================================================
# TAB 8 — ASSUMPTIONS & NOTES
# ============================================================
ws = wb.create_sheet("Assumptions & Notes")
ws.sheet_properties.tabColor = "777777"
ws.sheet_view.showGridLines = False

set_col_widths(ws, [4, 32, 60, 18])

title_block(
    ws,
    "Assumptions & Notes",
    "Transparent documentation of every assumption baked into this cost model.",
    last_col_letter="D",
)

header_row(ws, 4, ["#", "Assumption", "Reasoning / source", "Confidence"], height=30)

assumptions = [
    ("Volume baseline 200 candidates/day",
     "Industry-standard 20–40× funnel ratio to surface 8–15 qualified opportunities per week. "
     "Tunable; see Volume Sensitivity tab.", "High"),
    ("Twice-daily ingestion runs",
     "Captures intraday news / filings without overloading rate limits. Spec from NORAD blueprint.",
     "High"),
    ("~30% of candidates pass hard filters",
     "Typical drop rate after vertical/geo/size/blocked-tag deterministic filters. "
     "First-month real data will refine this.", "Medium"),
    ("~1.33% of processed candidates score ≥70",
     "80 qualified per 6,000 processed. Calibrated against benchmarks from analogous BD engines.",
     "Medium"),
    ("~0.2% of processed candidates score ≥85 (Elite)",
     "12 Elite per 6,000. Tunable threshold; if too tight, drops to 30/mo at 0.5%.",
     "Medium"),
    ("Diffbot ~3 credits per candidate",
     "1 KG lookup + 1 article extract + 1 NLP entities. Verified against Diffbot docs.",
     "High"),
    ("EXA ~5 searches + 10 contents per candidate",
     "Founders, funding, news, similar companies, niche industry — plus content for top hits.",
     "High"),
    ("Apify LinkedIn ~3 profiles per candidate",
     "1 company + 1–2 founders. Some candidates need none, some need more.",
     "Medium"),
    ("Apify PitchBook ~30% of candidates",
     "Only fires for private targets where funding signal matters. 1,800/mo.",
     "Medium"),
    ("Claude Sonnet 4.6 justification ~$0.039/call",
     "8K input + 1K output tokens at $3 in / $15 out per 1M.",
     "High"),
    ("Phase 2: 20 analyst seats × 5 NLS queries/day × 22 workdays",
     "Reasonable estimate for an active analyst team. Scales linearly if team grows.",
     "Medium"),
    ("Per NLS query ~$0.05 average",
     "Mix of cheap pgvector hits + Perplexity Sonar Pro + EXA + Claude synthesis.",
     "Medium"),
    ("Infra scales sub-linearly (~30% per 2× volume)",
     "Most infra cost is fixed (Supabase Team base, Langfuse Pro, Meilisearch Pro, Vercel seats). "
     "Variable: Fly machine count, Inngest executions, R2 storage.",
     "Medium"),
    ("Premium stack philosophy",
     "Per project requirements: pick the right tool per job, not the cheapest. "
     "Cheap-stack alternatives noted in Tool Catalog 'Alternative considered' column.",
     "Confirmed"),
    ("USD pricing throughout",
     "All vendor pricing listed in USD. Conversion to other currencies handled at billing layer.",
     "Confirmed"),
    ("Pricing snapshot date: April–May 2026",
     "Vendor pricing changes frequently. Refresh this model quarterly. Notable recent changes: "
     "Reddit went contract-only commercial in 2024, X API moved to pay-per-use Feb 2026, "
     "EXA repriced search to $7/1K in March 2026.",
     "Confirmed"),
    ("Excluded from this model",
     "CRM seats (HubSpot/Salesforce), email-sending (SendGrid/Postmark), PDF/report-gen services, "
     "downstream sales tooling, custom domains/SSL beyond Vercel/Fly, one-time setup or consulting.",
     "Confirmed"),
    ("Compliance baseline",
     "Default model assumes SOC2 required (drives Supabase Team $599 + Upstash Prod Pack $200). "
     "Drop to Supabase Pro + skip Prod Pack saves ~$770/mo if compliance not needed.",
     "Tunable"),
    ("Single-tenant per deployment",
     "Cost model is per NORAD deployment. Phase 2 multi-tenant in shared Supabase via RLS adds "
     "~$200/mo compute add-on but otherwise reuses the same infra.",
     "Confirmed"),
    ("BD-team-hour value @ $200/hr",
     "Used for the per-hour-saved KPI. Adjust upward for senior analysts ($300+/hr).",
     "Tunable"),
]
r = 5
for i, (a, why, conf) in enumerate(assumptions, start=1):
    fill_ = FILL_ALT if r % 2 else None
    write_row(ws, r, [i, a, why, conf],
              font=F_BODY, fill_=fill_, align=A_TOP_L, border=box())
    cc = ws.cell(row=r, column=4)
    cc.font = F_BODY_B
    cc.alignment = A_CENTER
    if conf == "High" or conf == "Confirmed":
        cc.fill = FILL_GOOD
    elif conf == "Medium":
        cc.fill = FILL_WARN
    else:
        cc.fill = FILL_BAD
    ws.row_dimensions[r].height = max(30, 16 * (1 + len(why) // 70))
    r += 1

# Footer
r += 2
ws.merge_cells(f"A{r}:D{r}")
c = ws.cell(row=r, column=1,
            value="Prepared April 2026 by Growth Machine · Project: NORAD Discovery Engine · "
                  "All vendor pricing verified against official pages April–May 2026.")
c.font = F_NOTE
c.alignment = A_LEFT

ws.freeze_panes = "A5"


# ============================================================
# SAVE
# ============================================================
out = "NORAD_COST_REPORT.xlsx"
wb.save(out)
print(f"Wrote {out}")
print(f"  v1 monthly total : ${total_v1:,}")
print(f"  P2 monthly total : ${total_p2:,}")
print(f"  Tabs: {wb.sheetnames}")
